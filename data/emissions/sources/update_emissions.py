# This script is made to collect data from different xlsx sources and to convert it to csvs data used for emissions visualisations
# Data conversion is made by Žiga Zaplotnik.
# Data scraping part contains a rework from a script made by Štefan Baebler for Sledilnik/zdravniki/data
# Vladimir Nešković Kesma, with a help from GPT4, has adapted the original scraping script and integrated Žigas script in one.  

import re
import requests
import os
import pandas as pd
import glob
import string
import openpyxl as xl
import sys
import hashlib
import numpy as np
from bs4 import BeautifulSoup
from datetime import datetime
from scipy.interpolate import interp1d

CONFIG_FILE = "./data/emissions/sources/update_emissions.config"
INTERMEDIATE_LATEST="./data/emissions/sources/emissions_historical_latest.xlsx"

# Lines that follow are data scraping part 
###########################################

# Function to extract the newest_env_string from the main page
def get_newest_env_string(soup):
    newest_string = soup.find("div", class_="filessection")
    newest_string = newest_string.find("a")["href"]
    return newest_string

# Function that returns the newest date
def get_newest_date(soup):
    # Use a regular expression to match the date format "14 Apr 2023"
    date_pattern = re.compile(r'\d{1,2} \w{3} \d{4}')

    # Find all <td> elements with class "tcenter"
    date_elements = soup.find_all('td', class_='tcenter')
    
    for date_element in date_elements:
        # Check if the class of the element is exactly "tcenter" and not more
        if ' '.join(date_element['class']) == 'tcenter':
            date_text = date_element.text.strip()
            # Check if the text matches the date format
            if date_pattern.match(date_text):
                # Parse the date text into a datetime object
                date_object = datetime.strptime(date_text, "%d %b %Y")
                # Format the datetime object into the desired format YYYYMMDD
                formatted_date = date_object.strftime("%Y%m%d")
                return formatted_date

    # Return an appropriate message or value if a valid date isn't found
    return "Valid date not found"


# Function to retrieve xlsx files from a given URL
def get_xlsx_files_from_url(url):
    page = requests.get(url)
    page.raise_for_status()
    soup = BeautifulSoup(page.content, "html.parser")
    em_elements = soup.find_all("em")
    filenames = [remove_html_tags(str(em)) for em in em_elements]
    return filter_strings_by_pattern(filter_xlsx_files(filenames))

# Function to remove HTML tags from a string
def remove_html_tags(input_string):
    return re.sub(r'<.*?>', '', input_string)

# Function to remove spaces and brackets from a list of strings
def remove_spaces_and_brackets_from_list(input_list):
    cleaned_list = [re.sub(r'[\[\]\s]', '', item) for item in input_list]
    return [item for item in cleaned_list if item]

# Function to filter xlsx files from a list of filenames
def filter_xlsx_files(file_list):
    return [filename for filename in file_list if filename.lower().endswith('.xlsx')]

# Function to filter strings by a specific pattern
def filter_strings_by_pattern(input_list):
    pattern = r'SVN_\d{4}_\d{4}.*'
    return [item for item in input_list if re.match(pattern, item)]

# Function to remove the last forward slash from a string
def remove_last_forward_slash(input_string):
    return input_string.rstrip("/")

# Function to read strings from a file and return them as a list
def read_strings_from_file(filename):
    with open(filename, 'r') as file:
        strings = file.read().splitlines()
    return strings

# Function to write a list of strings to a file
def write_strings_to_file(filename, strings):
    with open(filename, 'w') as file:
        for string in strings:
            file.write(string + '\n')

### Compute the SHA-1 checksum of a file
def compute_sha1_checksum(file_path):
    sha1 = hashlib.sha1()
    with open(file_path, 'rb') as file:
        # Read and update hash string value in blocks of 4K
        for byte_block in iter(lambda: file.read(4096), b""):
            sha1.update(byte_block)
    return sha1.hexdigest()

# count files in a directory
def count_files_in_directory(directory):
    # List all entries in the given directory
    entries = os.listdir(directory)
    # Filter out and count only files (not directories)
    file_count = sum(os.path.isfile(os.path.join(directory, entry)) for entry in entries)
    return file_count


# next two functions will read and write parameter vaules form/in a file
# parameters and values are separated by = and each new parameter with a new line
            
def readConfig(parameter, filename):
    try:
        with open(filename, 'r') as file:
            for line in file:
                key, value = line.strip().split('=')
                if key == parameter:
                    return value
        print(f"Parameter '{parameter}' not found.")
        sys.exit(1)
    except FileNotFoundError:
        print(f"File '{filename}' not found.")
        sys.exit(1)

def writeConfig(parameter, value, filename):
    try:
        # Read all lines from the file
        with open(filename, 'r') as file:
            lines = file.readlines()
        
        # Check if the parameter exists
        parameter_exists = False
        with open(filename, 'w') as file:
            for line in lines:
                key, _ = line.strip().split('=')
                if key == parameter:
                    line = f"{parameter}={value}\n"
                    parameter_exists = True
                file.write(line)
            
            # If the parameter does not exist, append it to the file
            if not parameter_exists:
                file.write(f"{parameter}={value}\n")
                
    except FileNotFoundError:
        print(f"File '{filename}' not found.")
        sys.exit(1)


# first of the main functions which to download emission xlsx files
###################################################################
def download_emission_xlsx_files():
    # URL where xlss are looked for
    BaseURL = "https://cdr.eionet.europa.eu/si/eu/mmr/art07_inventory/ghg_inventory/"
    # URL below will beceome new ULR from 15.3.2024 on
    # BaseURL = "https://cdr.eionet.europa.eu/si/eu/govreg/inventory/"
    page = requests.get(BaseURL)
    page.raise_for_status()
    soupMain = BeautifulSoup(page.content, "html.parser")
    

    # get the newest date
    newest_date = get_newest_date(soupMain)
    
    # Get the newest_env_string from the main page
    newest_env_string = get_newest_env_string(soupMain)

    # the real env string is without /
    if newest_env_string:
        addToBaseURL = remove_last_forward_slash(newest_env_string)
    
    new_env = readConfig("EU_LATEST_ENV",CONFIG_FILE) 

    #check if the download was already performed
    if new_env == addToBaseURL:
        print(f"There were no new files uploaded on {BaseURL}, continuing with the rest of the data processes")

    # if it was not, download the files
    else:

        # Generate URLs for the target pages (page 1 to 3)
        urls = [f"{BaseURL}{addToBaseURL}/index_html?&page={page_num}" for page_num in range(1, 4)]
        
        all_xlsx_files = []
        for url in urls:
            xlsx_files = get_xlsx_files_from_url(url)
            all_xlsx_files.extend(xlsx_files)
        
        #  download emission xlsx files and print the results

        for item in all_xlsx_files:
            #   ******* change this line ********
            new_date = readConfig("EU_LATEST_DATE",CONFIG_FILE)
            dest_dir = "./data/emissions/sources/" + newest_date
            dest = os.path.join(dest_dir, item)
        
            
            if  new_date == newest_date:
                print(f"    Already downloaded: {dest}")
            
            else:
                # calculating full URL 
                fullURL = BaseURL + addToBaseURL + "/" + item
                print(fullURL)
                r = requests.get(fullURL, allow_redirects=True)
                r.raise_for_status()
                # some spare parts for later
                # ct = r.headers.get('content-type')
                # if ct.lower() != "application/xlsx":
                #     print(f"Unexpected content type '{ct}'.")
                #     raise
                print(f"    Saving to: {dest}")
                directory = os.path.dirname(dest)
                # print(f"temp directory is {directory}")
                if not os.path.exists(directory):
                     os.makedirs(directory, exist_ok=True)
                open(dest, 'wb').write(r.content)
        writeConfig("EU_LATEST_DATE",newest_date,CONFIG_FILE)
        writeConfig("EU_LATEST_ENV",addToBaseURL,CONFIG_FILE)

# function to touch a file if it does not exists
        
def check_and_touch(file_path):
    # Check if the file exists
    if not os.path.exists(file_path):
        # Touch the file (create a new file) if it does not exist
        with open(file_path, 'a') as file:
            pass  # 'pass' is just a placeholder since no action is needed




# Lines that follow are for a method made by Žiga Zaplotnik and adapted a bit by Kesma, to integrate it with other part of the main scrip 
# method processes data for the files from the EU site above in an intermediate xlsx file used for other data processing later on           

def create_intermediate_xlsx():

    if compute_sha1_checksum(INTERMEDIATE_LATEST) == readConfig("CHKSUM_LATEST_INTERMEDIATE", CONFIG_FILE):
        print(f"{INTERMEDIATE_LATEST} already processed")
    else:
        inter_latest_date = readConfig("EU_LATEST_DATE",CONFIG_FILE)
        latest_dir = "./data/emissions/sources/" + inter_latest_date
        counted_files = count_files_in_directory(latest_dir)

        y = 1
        end_range = 1986 + counted_files 
        for year in range(1986,end_range):
            print(year)
            # Use a wildcard to match files with similar names

            file_pattern = "./data/emissions/sources/" + inter_latest_date + "/SVN_"+ inter_latest_date[:4] + "_{:04d}_*.xlsx".format(year)

            # Use the glob module to find all files that match the pattern
            files = glob.glob(file_pattern)
            print(files)
            # Loop through each file and read in the data using pandas
            for file in files:
                print(file)
                wb1 = xl.load_workbook(file)
                ws1 = wb1.worksheets[56]
        
            print (ws1)

            # opening the destination excel file 
            
            wb2 = xl.load_workbook(INTERMEDIATE_LATEST)
            ws2 = wb2.active

        
            # calculate total number of rows and 
            # columns in source excel file
            mr = ws1.max_row
            mc = ws1.max_column
            
            # copying the cell values from source 
            # excel file to destination excel file
            for i in range (1, mr + 1):
                # reading cell value from source excel file
                c = ws1.cell(row = 6+i, column = 10)
        
                # writing the read value to destination excel file
                # print(c.value)
                if isinstance(c.value, (int, float)):
                     ws2.cell(row=i+1, column=y+1).value = c.value
                else:
                    # Leave the cell empty
                    ws2.cell(row=i+1, column=y+1).value = None
                
                #ws2.cell(row = i+1, column = y+1).value = c.value
            
            # saving the destination excel file
            
            wb2.save(INTERMEDIATE_LATEST)
            
            y += 1
            writeConfig("CHKSUM_LATEST_INTERMEDIATE",compute_sha1_checksum(INTERMEDIATE_LATEST),CONFIG_FILE)


# Next lines of code are a method made by Žiga Zaplotnik and adapted a bit by Kesma, to integrate it with other part of the main scrip 
# method processes data from the historical transport and the intermediate data, processed from the EU site and creates a csv 

def transport_historical():
    TRANSPORT_LATEST="./data/emissions/sources/Emisije_TGP_iz_cestnega_prometa.xlsx"

    if compute_sha1_checksum(TRANSPORT_LATEST) == readConfig("CHKSUM_LATEST_TRANSPORT", CONFIG_FILE):
        print(f"{TRANSPORT_LATEST} already processed")
    else:
        # inter_latest_date = readConfig("EU_LATEST_DATE",CONFIG_FILE)
        # latest_dir = "./data/emissions/sources/" + inter_latest_date
        # counted_files = count_files_in_directory(latest_dir)

        # global warming potential of gases
        gwp = {"co2": 1,
                "ch4": 25,
                "n2o": 298,
                "nox": 0.,
                "co": 1.9,
                "nmvoc": 3.4,
                "so2": 0.,
                "cfc_hfc" : 1.}

        inter_latest_date = readConfig("EU_LATEST_DATE",CONFIG_FILE)
        
        year_start = 1986
        year_end = int(inter_latest_date[:4])-2
        print(year_end)

        years = np.arange(year_start,year_end+1)
        n = years.shape[0]

        transport__total = np.zeros(n)
        transport__domestic_aviation = np.zeros(n)
        transport__road_transporation = np.zeros(n)
        transport__road_transporation__cars = np.zeros(n)
        transport__road_transporation__light_duty_trucks = np.zeros(n)
        transport__road_transporation__heavy_duty_trucks = np.zeros(n)
        transport__road_transporation__buses = np.zeros(n)
        # transport__road_transporation__heavy_duty_trucks_buses = np.zeros(n)
        transport__road_transporation__motorcycles = np.zeros(n)
        transport__road_transporation__other = np.zeros(n)
        transport__railways = np.zeros(n)
        transport__domestic_navigation = np.zeros(n)
        transport__other_transportation = np.zeros(n)
        transport__international_aviation = np.zeros(n)
        transport__international_navigation = np.zeros(n)

        def co2equiv(arr,gwp):
            return arr[0]*gwp["co2"] + arr[1]*gwp["ch4"] + arr[2]*gwp["n2o"]

        def check_matrix(arr):
            if all([type(x)==str for x in arr]):
                return np.zeros(3)
            else:
                return arr


        def check_val(x):
            if type(x)==str:
                return 0
            else:
                return x

        i = 0
        y = year_start
        while y <= year_end:
            print(y)
            # manually loop through sheets
            # original
            # #file_name = glob.glob("./data/emissions/sources/20230414/SVN_{0:04d}_{1:04d}_*.xlsx".format(year_end,y))[0]
            # file_name = glob.glob("./data/emissions/sources/20230414/SVN_{0:04d}_{1:04d}_*.xlsx".format(year_end,y))
            # print (file_name)
            # sys.exit(0)

            #### VVVVVVVVVV NEW VVVVVVVVVVVV ######
            inter_latest_date = readConfig("EU_LATEST_DATE",CONFIG_FILE)
            file_pattern = "./data/emissions/sources/" + inter_latest_date + "/SVN_"+ inter_latest_date[:4] + "_{:04d}_*.xlsx".format(y)

            # Use the glob module to find all files that match the pattern
            file_name = glob.glob(file_pattern)[0]
            print(file_name)
            #### ^^^^^^^^^^^ NEW ^^^^^^^^^^^ ######
                
            # Sectoral report for energy, sheet 1 + sheet2
            df = pd.read_excel(file_name, sheet_name="Table1.A(a)s3")
            matrix = df.to_numpy()[7:-2,-3:]
            
            transport__total[i] = co2equiv(matrix[0],gwp)
            transport__domestic_aviation[i] = co2equiv(matrix[6],gwp)
            transport__road_transporation[i] = co2equiv(matrix[10],gwp)
            transport__road_transporation__cars[i] = co2equiv(matrix[18],gwp)
            transport__road_transporation__light_duty_trucks[i] = co2equiv(matrix[26],gwp)
            # transport__road_transporation__heavy_duty_trucks_buses[i] = co2equiv(matrix[34],gwp)
            transport__road_transporation__motorcycles[i] = co2equiv(matrix[42],gwp)
            transport__road_transporation__other[i] = co2equiv(check_matrix(matrix[51]),gwp)
            transport__railways[i] = co2equiv(matrix[52],gwp)
            transport__domestic_navigation[i] = co2equiv(check_matrix(matrix[58]),gwp)
            transport__other_transportation[i] = co2equiv(check_matrix(matrix[66]),gwp)
            
                            
            # add international aviation and navigation
            df = pd.read_excel(file_name,sheet_name="Summary2")
            matrix = df.to_numpy()
            
            transport__international_aviation[i] = matrix[56,-1]
            transport__international_navigation[i] = check_val(matrix[57,-1])
            
            
            # Division between  heavy duty trucks and buses
            file_name = glob.glob(TRANSPORT_LATEST)[0]
            df = pd.read_excel(file_name)
            matrix = df.to_numpy()[2:,3:]


            transport__road_transporation__heavy_duty_trucks[i] = co2equiv(matrix[::2,i],gwp)
            transport__road_transporation__buses[i] =  co2equiv(matrix[1::2,i],gwp)   
            
            y += 1
            i += 1

        df = pd.DataFrame({
                'year': years.astype(int),
                'total':     transport__total,
                'road_transporation.total': transport__road_transporation,
                'road_transporation.cars':    transport__road_transporation__cars,
                "road_transporation.light_duty_trucks":    transport__road_transporation__light_duty_trucks,
                'road_transporation.heavy_duty_trucks':     transport__road_transporation__heavy_duty_trucks,
                'road_transporation.buses':     transport__road_transporation__buses,
                'road_transporation.motorcycles': transport__road_transporation__motorcycles,
                'road_transporation.other': transport__road_transporation__other,
                'railways': transport__railways,
                'domestic_aviation': transport__domestic_aviation,
                'domestic_navigation': transport__domestic_navigation,
                'other_transportation': transport__other_transportation,
                'international_aviation': transport__international_aviation,
                'international_navigation': transport__international_navigation
                })
        df.to_csv("./data/emissions/data/emissions.historical.energy.transport.csv",index=False,float_format='%.2f')
        writeConfig("CHKSUM_LATEST_TRANSPORT",compute_sha1_checksum(TRANSPORT_LATEST),CONFIG_FILE)


# Next lines of code are a method made by Žiga Zaplotnik and adapted a bit by Kesma, to integrate it with other part of the main scrip 
# method processes data from the the intermediate data, processed from the EU site, and creates 6 csv-s

def emissions_historical():
    HISTORICAL_LATEST = "./data/emissions/sources/emissions_historical_latest.xlsx"
    check_and_touch(HISTORICAL_LATEST)
    if compute_sha1_checksum(HISTORICAL_LATEST) == readConfig("CHKSUM_LATEST_HISTORICAL", CONFIG_FILE):
        print(f"{HISTORICAL_LATEST} already processed for historical csvs")
    else:
        df = pd.read_excel(INTERMEDIATE_LATEST,sheet_name="Sheet1")
        emissions_historical = df.values[:,1:]

        # emission indices 
        ei= {"total_net" : 0 ,
            "energy" : {
                    "total" : 1, 
                    "fuel_combustion_activities" : {
                            "total" : 2,
                            "energy_industries" : 3,
                            "manufacturing_construction" : 4,
                            "transport" : 5,
                            "other_sectors" : 6,
                            "other" : 7                            
                            },
                    "fugitive_emissions_from_fuels" : {
                            "total" : 8,
                            "solid_fuels" : 9,
                            "oil_natural_gas_and_energy_production" : 10
                            },
                    "co2_transport_storage" : 11
                    },
            "industrial_processes" : {
                    "total" : 12,
                    "mineral_industry" : 13,
                    "chemical_industry": 14,
                    "metal_industry" : 15,
                    "non_energy_products_from_fuels" : 16,
                    "electronic_industry" : 17,
                    "product_usese_as_ODS" : 18,
                    "other_product_manufacture_use":19,
                    "other": 20
                    },
            "agriculture" : {
                    "total": 21,
                    "enteric_fermentation": 22,
                    "manure_management" : 23,
                    "rice_cultivation" : 24,
                    "agricultural_soils" : 25,
                    "prescribed_burning_of_savannas" : 26,
                    "field_burning_agricultural_residues" : 27,
                    "liming" : 28,
                    "urea_application" : 29,
                    "carbon_containing_fertilizers" : 30,
                    "other" : 31
                    },
            "lulucf" : {
                    "total" : 32,
                    "forest_land" : 33,
                    "cropland" : 34,
                    "grassland" : 35,
                    "wetlands": 36,
                    "settlements": 37,
                    "other_land": 38,
                    "harvested_wood_prducts":39,
                    "other" : 40
                    },
            "waste": {
                    "total" : 41,
                    "solid_waste_disposal" : 42,
                    "biological_treatment_solid_waste" : 43,
                    "incineration_open_burning_waste" : 44,
                    "waste_water_treatment_discharge" : 45,
                    "other" : 46
                    },
            "other" : 47,
            "international_bunkers": {
                    "total": 50,
                    "aviation" : 51,
                    "navigation" : 52
                    },
            "multilateral_operations" : 53,
            "co2_emissions_from_biomass" : 54,
            "co2_captured" : 55,
            "longerim_storage_waste_disposal" : 56,
            "indirect_n20" : 57,
            "indirect_co2" : 58,
            "total_source" : 59,
            }
            
        historical_latest_date = readConfig("EU_LATEST_DATE",CONFIG_FILE)
    
        # not sure if this formula is OK
        years = np.arange(1986,int(historical_latest_date[:4]) - 1)

        print(df)

        df = pd.DataFrame({
                'year': years.astype(int),
                'total_source':                     emissions_historical[ei["total_source"]],
                'energy_industries':                emissions_historical[ei["energy"]["fuel_combustion_activities"]["energy_industries"]],
                'manufacturing_construction_fuels': emissions_historical[ei["energy"]["fuel_combustion_activities"]["manufacturing_construction"]],
                'transport':                        emissions_historical[ei["energy"]["fuel_combustion_activities"]["transport"]],
                "industrial_processes":             emissions_historical[ei["industrial_processes"]["total"]],
                'residential_commercial_agricultural_forestry_fishing_fuels':  emissions_historical[ei["energy"]["fuel_combustion_activities"]["other_sectors"]],
                'agriculture':                      emissions_historical[ei["agriculture"]["total"]],
                'waste':                            emissions_historical[ei["waste"]["total"]],
                "international_aviation":           emissions_historical[ei["international_bunkers"]["aviation"]],
                "international_navigation":         emissions_historical[ei["international_bunkers"]["navigation"]],
                'co2_emissions_from_biomass':       emissions_historical[ei["co2_emissions_from_biomass"]],
                'others':                           emissions_historical[ei["energy"]["fuel_combustion_activities"]["other"]] +\
                                                    emissions_historical[ei["energy"]["fugitive_emissions_from_fuels"]["total"]],
                'lulucf':                           emissions_historical[ei["lulucf"]["total"]]
                })
        df.to_csv("./data/emissions/data/emissions.historical.csv",index=False,float_format='%.2f')
        


        df = pd.DataFrame({
                'year': years.astype(int),
                'total':                                                                emissions_historical[ei["energy"]["total"]],
                'fuel_combustion_activities.total':                                     emissions_historical[ei["energy"]["fuel_combustion_activities"]["total"]],
                'fuel_combustion_activities.energy_industries':                         emissions_historical[ei["energy"]["fuel_combustion_activities"]["energy_industries"]],
                'fuel_combustion_activities.manufacturing_construction':                emissions_historical[ei["energy"]["fuel_combustion_activities"]["manufacturing_construction"]],
                'fuel_combustion_activities.transport':                                 emissions_historical[ei["energy"]["fuel_combustion_activities"]["transport"]],
                'fuel_combustion_activities.other_sectors':                             emissions_historical[ei["energy"]["fuel_combustion_activities"]["other_sectors"]],
                'fuel_combustion_activities.other':                                     emissions_historical[ei["energy"]["fuel_combustion_activities"]["other"]],
                'fugitive_emissions_from_fuels.total':                                  emissions_historical[ei["energy"]["fugitive_emissions_from_fuels"]["total"]],
                'fugitive_emissions_from_fuels.solid_fuels':                            emissions_historical[ei["energy"]["fugitive_emissions_from_fuels"]["solid_fuels"]],
                'fugitive_emissions_from_fuels.oil_natural_gas_and_energy_production':  emissions_historical[ei["energy"]["fugitive_emissions_from_fuels"]["oil_natural_gas_and_energy_production"]],
                'co2_transport_storage':                                                emissions_historical[ei["energy"]["co2_transport_storage"]]
                })
        df.to_csv("./data/emissions/data/emissions.historical.energy.csv",index=False,float_format='%.2f')
        

        df = pd.DataFrame({
                'year': years.astype(int),
                'total':                emissions_historical[ei["industrial_processes"]["total"]],
                'mineral_industry': emissions_historical[ei["industrial_processes"]["mineral_industry"]],
                'chemical_industry':    emissions_historical[ei["industrial_processes"]["chemical_industry"]],
                'metal_industry':     emissions_historical[ei["industrial_processes"]["metal_industry"]],
                'non_energy_products_from_fuels': emissions_historical[ei["industrial_processes"]["non_energy_products_from_fuels"]],
                'electronic_industry': emissions_historical[ei["industrial_processes"]["electronic_industry"]],
                'product_usese_as_ODS': emissions_historical[ei["industrial_processes"]["product_usese_as_ODS"]],
                'other_product_manufacture_use': emissions_historical[ei["industrial_processes"]["other_product_manufacture_use"]],
                'other': emissions_historical[ei["industrial_processes"]["other"]]
                })
        df.to_csv("./data/emissions/data/emissions.historical.industrial.processes.csv",index=False,float_format='%.2f')
        

        df = pd.DataFrame({
                'year': years.astype(int),
                'total':                emissions_historical[ei["agriculture"]["total"]],
                'enteric_fermentation': emissions_historical[ei["agriculture"]["enteric_fermentation"]],
                'manure_management':    emissions_historical[ei["agriculture"]["manure_management"]],
                'rice_cultivation':     emissions_historical[ei["agriculture"]["rice_cultivation"]],
                'agricultural_soils':     emissions_historical[ei["agriculture"]["agricultural_soils"]],
                'prescribed_burning_of_savannas': emissions_historical[ei["agriculture"]["prescribed_burning_of_savannas"]],
                'field_burning_agricultural_residues': emissions_historical[ei["agriculture"]["field_burning_agricultural_residues"]],
                'liming': emissions_historical[ei["agriculture"]["liming"]],
                'urea_application': emissions_historical[ei["agriculture"]["urea_application"]],
                'carbon_containing_fertilizers': emissions_historical[ei["agriculture"]["carbon_containing_fertilizers"]],
                'other':                emissions_historical[ei["agriculture"]["other"]]
                })
        df.to_csv("./data/emissions/data/emissions.historical.agriculture.csv",index=False,float_format='%.2f')


        df = pd.DataFrame({
                'year': years.astype(int),
                'total':                emissions_historical[ei["lulucf"]["total"]],
                'forest_land': emissions_historical[ei["lulucf"]["forest_land"]],
                'cropland':    emissions_historical[ei["lulucf"]["cropland"]],
                "grassland":    emissions_historical[ei["lulucf"]["grassland"]],
                'wetlands':     emissions_historical[ei["lulucf"]["wetlands"]],
                'settlements': emissions_historical[ei["lulucf"]["settlements"]],
                'other_land': emissions_historical[ei["lulucf"]["other_land"]],
                'harvested_wood_prducts': emissions_historical[ei["lulucf"]["harvested_wood_prducts"]],
                'other': emissions_historical[ei["lulucf"]["other"]]
                })
        df.to_csv("./data/emissions/data/emissions.historical.lulucf.csv",index=False,float_format='%.2f')


        df = pd.DataFrame({
                'year': years.astype(int),
                'total':                emissions_historical[ei["waste"]["total"]],
                'solid_waste_disposal': emissions_historical[ei["waste"]["solid_waste_disposal"]],
                'biological_treatment_solid_waste':    emissions_historical[ei["waste"]["biological_treatment_solid_waste"]],
                'incineration_open_burning_waste':    emissions_historical[ei["waste"]["incineration_open_burning_waste"]],
                "waste_water_treatment_discharge":    emissions_historical[ei["waste"]["waste_water_treatment_discharge"]],
                'other':     emissions_historical[ei["waste"]["other"]]
                })
        df.to_csv("./data/emissions/data/emissions.historical.waste.csv",index=False,float_format='%.2f')


        df = pd.DataFrame({
                'year': years.astype(int),
                'international_bunkers.total':      emissions_historical[ei["international_bunkers"]["total"]],
                'international_bunkers.aviation':   emissions_historical[ei["international_bunkers"]["aviation"]],
                'international_bunkers.navigation': emissions_historical[ei["international_bunkers"]["navigation"]],
                'multilateral_operations':          emissions_historical[ei["multilateral_operations"]],
                'co2_emissions_from_biomass':       emissions_historical[ei["co2_emissions_from_biomass"]],
                'co2_captured':    emissions_historical[ei["co2_captured"]],
                "longerim_storage_waste_disposal":    emissions_historical[ei["longerim_storage_waste_disposal"]],
                "indirect_n20":    emissions_historical[ei["indirect_n20"]],
                "indirect_co2":    emissions_historical[ei["indirect_co2"]]
                })
        df.to_csv("./data/emissions/data/emissions.historical.memo_items.csv",index=False,float_format='%.2f')

        writeConfig("CHKSUM_LATEST_HISTORICAL",compute_sha1_checksum(HISTORICAL_LATEST),CONFIG_FILE)

# Next lines of code are a method made by Žiga Zaplotnik and adapted a bit by Kesma, to integrate it with other part of the main script 
# method processes data from the the paris projections data and creates 1 csv 

def paris_projections():
    PARIS_PROJECTIONS = "./data/emissions/sources/ProjekcijeGHG_Slovenija.xlsx"
    if compute_sha1_checksum(PARIS_PROJECTIONS) == readConfig("CHKSUM_PARIS", CONFIG_FILE):
        print(f"{PARIS_PROJECTIONS} already processed")
    else:
        years = np.arange(2020,2031)
        df = pd.read_excel(PARIS_PROJECTIONS)
        matrix = df.to_numpy()
        bau = matrix[:11,3]
        nepn = matrix[:11,4]
        # print (nepn)
        # raise SystemExit(0)
        ec = matrix[:11,6]
        paris20 = matrix[:11,8]
        paris15 = matrix[:11,10]

        df = pd.DataFrame({'year': years.astype(int),
                        'bau': bau,
                        'nepn':nepn,
                        'ec':ec,
                        'paris20':paris20,
                        'paris15':paris15})
        df.to_csv("./data/emissions/data/emissions.projections.ec_paris.csv",index=False,float_format="%.2f")
        writeConfig("CHKSUM_PARIS",compute_sha1_checksum(PARIS_PROJECTIONS),CONFIG_FILE)

# Next lines of code are a method made by Žiga Zaplotnik and adapted a bit by Kesma, to integrate it with other part of the main script 
# method processes data from the the NEPN projections data and creates 6 csvs

def nepn_projections():
    NEPN_PROJECTIONS = "./data/emissions/sources/energetska_bilanca_2050_nepn_dps.xlsx"
    if compute_sha1_checksum(NEPN_PROJECTIONS) == readConfig("CHKSUM_NEPN", CONFIG_FILE):
        print(f"{NEPN_PROJECTIONS} already processed")
    else:
        sp_years = np.arange(2020,2051,5)

        df = pd.read_excel(NEPN_PROJECTIONS,sheet_name="EmisijeTGP")

        # current measures
        sp_emissions_projections_current = df.to_numpy()[23:44,20:27]

        # business as usual
        sp_emissions_projections_bau = df.to_numpy()[23:44,60:67]

        # additional measures, nuclear
        sp_emissions_projections_add_nuc = df.to_numpy()[23:44,28:35]

        # additional measures, synthetic
        sp_emissions_projections_add_syn = df.to_numpy()[23:44,36:43]

        # ambitious additional measures, nuclear
        sp_emissions_projections_ambadd_nuc = df.to_numpy()[23:44,44:51]

        # ambitious additional measures, synthetic
        sp_emissions_projections_ambadd_syn = df.to_numpy()[23:44,52:59]


        #%% perform linear interpolation
        years = np.arange(2020,2051,1)

        fun_emissions_projections_current = interp1d(sp_years,sp_emissions_projections_current,axis=1)
        emissions_projections_current = fun_emissions_projections_current(years)

        fun_emissions_projections_bau = interp1d(sp_years,sp_emissions_projections_bau,axis=1)
        emissions_projections_bau = fun_emissions_projections_bau(years)

        fun_emissions_projections_add_nuc = interp1d(sp_years,sp_emissions_projections_add_nuc,axis=1)
        emissions_projections_add_nuc = fun_emissions_projections_add_nuc(years)

        fun_emissions_projections_add_syn = interp1d(sp_years,sp_emissions_projections_add_syn,axis=1)
        emissions_projections_add_syn = fun_emissions_projections_add_syn(years)

        fun_emissions_projections_ambadd_nuc = interp1d(sp_years,sp_emissions_projections_ambadd_nuc,axis=1)
        emissions_projections_ambadd_nuc = fun_emissions_projections_ambadd_nuc(years)

        fun_emissions_projections_ambadd_syn = interp1d(sp_years,sp_emissions_projections_ambadd_syn,axis=1)
        emissions_projections_ambadd_syn = fun_emissions_projections_ambadd_syn(years)

        #%%
        # emission indices from file emissions_total.xlsx
        ei= {"energy" : {
                    "total" : 0, 
                    "fuel_combustion_activities" : {
                            "total" : 1,
                            "energy_industries" : 2,
                            "manufacturing_construction" : 3,
                            "transport" : 4,
                            "other_sectors" : 5,
                            "other" : 6                            
                            },
                    "fugitive_emissions_from_fuels" : {
                            "total" : 7,
                            "solid_fuels" : None,
                            "oil_natural_gas_and_energy_production" : None
                            },
                    "co2_transport_storage" : None
                    },
            "industrial_processes" : {
                    "total" : 8,
                    "mineral_industry" : None,
                    "chemical_industry": None,
                    "metal_industry" : None,
                    "non_energy_products_from_fuels" : None,
                    "electronic_industry" : None,
                    "product_usese_as_ODS" : None,
                    "other_product_manufacture_use": None,
                    "other": None
                    },
            "agriculture" : {
                    "total": 9,
                    "enteric_fermentation": None,
                    "manure_management" : None,
                    "rice_cultivation" : None,
                    "agricultural_soils" : None,
                    "prescribed_burning_of_savannas" : None,
                    "field_burning_agricultural_residues" : None,
                    "liming" : None,
                    "urea_application" : None,
                    "carbon_containing_fertilizers" : None,
                    "other" : None
                    },
            "lulucf" : {
                    "total" : 10,
                    "forest_land" : 11,
                    "cropland" : 12,
                    "grassland" : 13,
                    "wetlands": 14,
                    "settlements": 15,
                    "other_land": 16,
                    "harvested_wood_prducts":17,
                    "other" : None
                    },
            "waste": {
                    "total" : 18,
                    "solid_waste_disposal" : None,
                    "biological_treatment_solid_waste" : None,
                    "incineration_open_burning_waste" : None,
                    "waste_water_treatment_discharge" : None,
                    "other" : None
                    },
            "other" : None,
            "international_bunkers": {
                    "total": None,
                    "aviation" : None,
                    "navigation" : None
                    },
            "multilateral_operations" : None,
            "co2_emissions_from_biomass" : None,
            "co2_captured" : None,
            "longerim_storage_waste_disposal" : None,
            "indirect_n20" : None,
            "indirect_co2" : None,
            "total_source" : 19
            }
            
        #%% export data    
        def export_data(data,years,fnm):    
            df = pd.DataFrame({
                    'year': years.astype(int),
                    'total_source':                     data[ei["total_source"]],
                    'energy_industries':                data[ei["energy"]["fuel_combustion_activities"]["energy_industries"]],
                    'manufacturing_construction_fuels': data[ei["energy"]["fuel_combustion_activities"]["manufacturing_construction"]],
                    'transport':                        data[ei["energy"]["fuel_combustion_activities"]["transport"]],
                    "industrial_processes":             data[ei["industrial_processes"]["total"]],
                    'residential_commercial_agricultural_forestry_fishing_fuels':  data[ei["energy"]["fuel_combustion_activities"]["other_sectors"]],
                    'agriculture':                      data[ei["agriculture"]["total"]],
                    'waste':                            data[ei["waste"]["total"]],
                    'others':                           data[ei["energy"]["fuel_combustion_activities"]["other"]] +\
                                                        data[ei["energy"]["fugitive_emissions_from_fuels"]["total"]],
                    'lulucf':                           data[ei["lulucf"]["total"]]
                    })
            df.to_csv(fnm,index=False,float_format='%.2f')
        
            
        export_data(emissions_projections_current,years,"./data/emissions/data/emissions.projections.current.csv")
        export_data(emissions_projections_bau,years,"./data/emissions/data/emissions.projections.bau.csv")
        export_data(emissions_projections_add_nuc,years,"./data/emissions/data/emissions.projections.additional_nuclear.csv")
        export_data(emissions_projections_add_syn,years,"./data/emissions/data/emissions.projections.additional_synthetic.csv")
        export_data(emissions_projections_ambadd_nuc,years,"./data/emissions/data/emissions.projections.ambitious_additional_nuclear.csv")
        export_data(emissions_projections_ambadd_syn,years,"./data/emissions/data/emissions.projections.ambitious_additional_synthetic.csv")

        writeConfig("CHKSUM_NEPN",compute_sha1_checksum(NEPN_PROJECTIONS),CONFIG_FILE)


########## main ############
############################
        
if __name__ == "__main__":
    # Download emission xlsx files 
    download_emission_xlsx_files()

    # Create intermediate xlsx
    create_intermediate_xlsx()

    # process transport emissions waiting for the 
    transport_historical()

    # process historical emissions
    emissions_historical()

    # process paris projections
    paris_projections()

    # process NEPN projections
    nepn_projections()
